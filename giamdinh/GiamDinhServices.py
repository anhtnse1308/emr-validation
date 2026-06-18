"""
GiamDinhService – Orchestrator chạy toàn bộ rule giám định và xuất Excel.

Luồng:
  GiamDinhService.run(all_objects) → list[GiamDinhError]
  GiamDinhService.write_excel_giamdinh(excel_file, errors) → path file output

Thêm rule mới: import class, thêm vào RULES list – xong.
"""

import os
from datetime import datetime
from dataclasses import fields as dc_fields

from giamdinh.GiamDinhBase import GiamDinhError
from giamdinh.rules.ChongChiDinh import ChongChiDinh
from giamdinh.rules.DichVuTrongQuyTrinh import DichVuTrongQuyTrinh
from giamdinh.rules.VTYTVuotMuc import VTYTVuotMuc
from giamdinh.rules.ThuocChiDinh import ThuocChiDinh
from giamdinh.rules.LuotKhamBenhNhan import LuotKhamBenhNhan
from giamdinh.rules.GioiHanTanSuat import GioiHanTanSuat
from giamdinh.rules.DieuKienDacBiet import DieuKienDacBiet
from giamdinh.rules.KiemTraDuLieu import KiemTraDuLieu


class GiamDinhService:
    """Chạy toàn bộ rule giám định trên all_objects."""

    # Đăng ký rule tại đây – thứ tự = thứ tự xuất hiện trong báo cáo
    RULES = [
        ChongChiDinh(),
        DichVuTrongQuyTrinh(),
        VTYTVuotMuc(),
        ThuocChiDinh(),
        LuotKhamBenhNhan(),
        GioiHanTanSuat(),
        DieuKienDacBiet(),
        KiemTraDuLieu(),
    ]

    @classmethod
    def run(cls, all_objects: dict) -> list[GiamDinhError]:
        """Chạy tất cả rule, trả về danh sách lỗi giám định."""
        all_errors: list[GiamDinhError] = []
        for rule in cls.RULES:
            try:
                result = rule.check(all_objects)
                all_errors.extend(result)
            except Exception as e:
                # Không để 1 rule lỗi phá toàn bộ luồng
                print(f"[GIAMDINH] Lỗi rule {rule.__class__.__name__}: {e}")
        return all_errors

    # -------------------------------------------------------------------
    # Xuất Excel giám định
    # -------------------------------------------------------------------
    @staticmethod
    def write_excel_giamdinh(excel_file: str, errors: list[GiamDinhError]) -> str:
        """
        Xuất kết quả giám định ra file Excel.
        Tên file: <ten_goc>[<timestamp>]_giamdinh.xlsx

        Sheet 1 – TONG_HOP:
          Thống kê số lỗi theo sheet XML + mã lý do từ chối

        Sheet 2 – CHI_TIET:
          Toàn bộ GiamDinhError dạng bảng, đầy đủ thông tin

        Sheet per rule group (S011.1, S054.1, D004.1, S059.2, S079.1):
          Lọc riêng từng nhóm lý do để dễ tra cứu
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        except ImportError:
            raise ImportError("Cần cài openpyxl: pip install openpyxl")

        # -------------------------------------------------------------------
        # Style
        # -------------------------------------------------------------------
        FILL_HDR    = PatternFill("solid", start_color="1F3864")   # xanh đậm
        FILL_HDR2   = PatternFill("solid", start_color="C00000")   # đỏ đậm (nhóm sheet)
        FILL_ROW_ODD  = PatternFill("solid", start_color="FFF2CC") # vàng nhạt
        FILL_ROW_EVEN = PatternFill("solid", start_color="FFFFFF") # trắng
        FILL_WARN   = PatternFill("solid", start_color="FCE4D6")   # cam nhạt (WARN)
        FILL_OK     = PatternFill("solid", start_color="E2EFDA")   # xanh nhạt (tổng hợp OK)

        FONT_HDR    = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        FONT_DATA   = Font(name="Arial", size=10)
        FONT_BOLD   = Font(bold=True, name="Arial", size=10)
        FONT_RED    = Font(name="Arial", size=10, color="C00000")
        FONT_WARN   = Font(name="Arial", size=10, color="833C00", italic=True)

        ALIGN_CTR   = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ALIGN_L     = Alignment(vertical="center", wrap_text=True)
        ALIGN_L_TOP = Alignment(vertical="top", wrap_text=True)

        _THIN  = Side(style="thin", color="AAAAAA")
        BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

        def _hdr(ws, row, col, val, fill=FILL_HDR):
            c = ws.cell(row=row, column=col, value=val)
            c.fill      = fill
            c.font      = FONT_HDR
            c.alignment = ALIGN_CTR
            c.border    = BORDER
            return c

        def _cell(ws, row, col, val, fill=FILL_ROW_EVEN,
                  font=FONT_DATA, align=ALIGN_L):
            c = ws.cell(row=row, column=col, value=val)
            c.fill      = fill
            c.font      = font
            c.alignment = align
            c.border    = BORDER
            return c

        def _auto_width(ws, headers, start_row=2):
            from openpyxl.utils import get_column_letter
            for ci, h in enumerate(headers, 1):
                max_w = len(str(h))
                for ri in range(start_row, ws.max_row + 1):
                    v = ws.cell(row=ri, column=ci).value or ""
                    line1 = str(v).split("\n")[0]
                    max_w = max(max_w, min(len(line1), 80))
                ws.column_dimensions[get_column_letter(ci)].width = max_w + 3

        # -------------------------------------------------------------------
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)

        now_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        # ===================================================================
        # Sheet 1 – TONG_HOP
        # ===================================================================
        ws_th = wb.create_sheet("TONG_HOP")

        # Tiêu đề
        ws_th.merge_cells("A1:F1")
        title_cell = ws_th.cell(row=1, column=1,
            value=f"TỔNG HỢP KẾT QUẢ GIÁM ĐỊNH BHYT  –  {now_str}")
        title_cell.fill      = FILL_HDR
        title_cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=12)
        title_cell.alignment = ALIGN_CTR
        ws_th.row_dimensions[1].height = 30

        # Header bảng tổng hợp
        th_headers = ["MÃ LÝ DO", "TÊN NHÓM LỖI", "SHEET XML", "SỐ LƯỢNG LỖI", "GHI CHÚ"]
        for ci, h in enumerate(th_headers, 1):
            _hdr(ws_th, 2, ci, h)

        # Tổng hợp theo mã lý do + sheet
        from collections import Counter
        counter: Counter = Counter()
        for err in errors:
            counter[(err.ma_ly_do, err.sheet)] += 1

        NHOM_MO_TA = {
            "S011.1":      "Dịch vụ kỹ thuật có chống chỉ định",
            "S054.1":      "DVKT nằm trong quy trình của DVKT khác",
            "D004.1":      "VTYT vượt mức chi quy định",
            "S059.2":      "Thuốc YHCT ngoài giới hạn chỉ định",
            "S079.1":      "Kháng sinh đặc biệt sai hạng BV",
            "S079.1-WARN": "Kháng sinh đặc biệt – CẦN KIỂM TRA THỦ CÔNG",
            "":            "Lỗi nghiệp vụ khác (không có mã từ chối riêng)",
        }

        ri = 3
        for (ma_ly_do, sheet), cnt in sorted(counter.items()):
            fill = FILL_WARN if "WARN" in (ma_ly_do or "") else FILL_ROW_ODD
            _cell(ws_th, ri, 1, ma_ly_do, fill=fill, font=FONT_BOLD)
            _cell(ws_th, ri, 2, NHOM_MO_TA.get(ma_ly_do, ""), fill=fill)
            _cell(ws_th, ri, 3, sheet, fill=fill, align=ALIGN_CTR)
            _cell(ws_th, ri, 4, cnt, fill=fill, font=FONT_RED, align=ALIGN_CTR)
            _cell(ws_th, ri, 5, "", fill=fill)
            ri += 1

        # Tổng cộng
        _cell(ws_th, ri, 1, "TỔNG", fill=FILL_HDR, font=FONT_HDR, align=ALIGN_CTR)
        _cell(ws_th, ri, 2, "", fill=FILL_HDR)
        _cell(ws_th, ri, 3, "", fill=FILL_HDR)
        _cell(ws_th, ri, 4, len(errors), fill=FILL_HDR, font=FONT_HDR, align=ALIGN_CTR)
        _cell(ws_th, ri, 5, "", fill=FILL_HDR)

        _auto_width(ws_th, th_headers)
        ws_th.freeze_panes = "A3"

        # ===================================================================
        # Sheet 2 – CHI_TIET (tất cả lỗi)
        # ===================================================================
        ws_ct = wb.create_sheet("CHI_TIET")
        ct_headers = [
            "STT", "SHEET", "ROW_EXCEL", "MA_LK",
            "MA_DICH_VU", "MA_LY_DO", "MO_TA", "CAN_CU",
        ]
        for ci, h in enumerate(ct_headers, 1):
            _hdr(ws_ct, 1, ci, h)

        for idx, err in enumerate(errors, 1):
            ri = idx + 1
            fill = FILL_WARN if "WARN" in (err.ma_ly_do or "") else (
                FILL_ROW_ODD if idx % 2 == 1 else FILL_ROW_EVEN
            )
            _cell(ws_ct, ri, 1, idx, fill=fill, align=ALIGN_CTR)
            _cell(ws_ct, ri, 2, err.sheet, fill=fill, align=ALIGN_CTR)
            _cell(ws_ct, ri, 3, err.row_excel, fill=fill, align=ALIGN_CTR)
            _cell(ws_ct, ri, 4, err.ma_lk, fill=fill)
            _cell(ws_ct, ri, 5, err.ma_dich_vu, fill=fill)
            _cell(ws_ct, ri, 6, err.ma_ly_do, fill=fill,
                  font=FONT_RED if err.ma_ly_do else FONT_DATA, align=ALIGN_CTR)
            _cell(ws_ct, ri, 7, err.mo_ta, fill=fill, align=ALIGN_L_TOP)
            _cell(ws_ct, ri, 8, err.can_cu, fill=fill, align=ALIGN_L_TOP)
            ws_ct.row_dimensions[ri].height = max(
                30, len(err.mo_ta) // 60 * 15 + 15
            )

        _auto_width(ws_ct, ct_headers)
        ws_ct.freeze_panes = "A2"
        ws_ct.row_dimensions[1].height = 30

        # ===================================================================
        # Sheet per nhóm mã lý do
        # ===================================================================
        ma_ly_do_groups: dict[str, list] = {}
        for err in errors:
            key = err.ma_ly_do or "KHAC"
            ma_ly_do_groups.setdefault(key, []).append(err)

        for ma_ly_do, group_errors in sorted(ma_ly_do_groups.items()):
            # Tên sheet tối đa 31 ký tự (giới hạn Excel)
            sheet_name = ma_ly_do.replace(".", "_").replace("-", "_")[:31]
            if sheet_name in [ws.title for ws in wb.worksheets]:
                sheet_name = sheet_name[:28] + "_2"

            ws_g = wb.create_sheet(sheet_name)
            g_headers = [
                "STT", "SHEET", "ROW_EXCEL", "MA_LK",
                "MA_DICH_VU", "MO_TA", "CAN_CU",
            ]
            for ci, h in enumerate(g_headers, 1):
                _hdr(ws_g, 1, ci, h, fill=FILL_HDR2)

            for idx, err in enumerate(group_errors, 1):
                ri = idx + 1
                fill = FILL_ROW_ODD if idx % 2 == 1 else FILL_ROW_EVEN
                _cell(ws_g, ri, 1, idx, fill=fill, align=ALIGN_CTR)
                _cell(ws_g, ri, 2, err.sheet, fill=fill, align=ALIGN_CTR)
                _cell(ws_g, ri, 3, err.row_excel, fill=fill, align=ALIGN_CTR)
                _cell(ws_g, ri, 4, err.ma_lk, fill=fill)
                _cell(ws_g, ri, 5, err.ma_dich_vu, fill=fill)
                _cell(ws_g, ri, 6, err.mo_ta, fill=fill, align=ALIGN_L_TOP)
                _cell(ws_g, ri, 7, err.can_cu, fill=fill, align=ALIGN_L_TOP)
                ws_g.row_dimensions[ri].height = max(
                    30, len(err.mo_ta) // 60 * 15 + 15
                )

            _auto_width(ws_g, g_headers)
            ws_g.freeze_panes = "A2"
            ws_g.row_dimensions[1].height = 30

        # ===================================================================
        # Lưu file
        # ===================================================================
        folder   = os.path.dirname(os.path.abspath(excel_file))
        base     = os.path.splitext(os.path.basename(excel_file))[0]
        now_ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_file = os.path.join(folder, f"{base}[{now_ts}]_giamdinh.xlsx")
        wb.save(out_file)
        return out_file