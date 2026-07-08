"""
Rule – Kiểm tra ngày giường bệnh nhân nội trú (MA_LOAI_KCB = 03)
Căn cứ: TT 39/2018/TT-BYT + TT 13/2019/TT-BYT

Detect từ XML1 + XML3:
  G1 – Số dịch vụ giường ≠ số ngày điều trị          → WARN.GG01
  G2 – Cùng MA_LK, cùng ngày có > 1 dịch vụ giường   → WARN.GG02

Logic:
  - "Dịch vụ giường" = dòng XML3 có TEN_DICH_VU chứa từ "giường" (không phân biệt hoa/thường)
  - Mỗi dòng giường = 1 ngày, xác định bởi NGAY_TH_YL[:8]
  - Số ngày điều trị = (NGAY_RA.date - NGAY_VAO.date).days + 1  (luôn dùng công thức này)
  - Chỉ áp dụng MA_LOAI_KCB = "03" (nội trú)
"""

from __future__ import annotations
import os
from collections import defaultdict
from datetime import datetime, date

from giamdinh.GiamDinhBase import GiamDinhBase, GiamDinhError

MA_LOAI_KCB_NOI_TRU = "03"
CAN_CU = "TT 39/2018/TT-BYT + TT 13/2019/TT-BYT (giá dịch vụ ngày giường)"


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------
def _parse_date(val: str) -> date | None:
    """Lấy phần ngày từ DATE12 hoặc DATE8."""
    s = str(val or "").strip()[:8]
    if len(s) == 8 and s.isdigit():
        try:
            return datetime.strptime(s, "%Y%m%d").date()
        except ValueError:
            pass
    return None


def _so_ngay(ngay_vao: str, ngay_ra: str) -> int | None:
    """Số ngày điều trị = NGAY_RA - NGAY_VAO + 1."""
    d_vao = _parse_date(ngay_vao)
    d_ra  = _parse_date(ngay_ra)
    if d_vao and d_ra and d_ra >= d_vao:
        return (d_ra - d_vao).days + 1
    return None


def _is_giuong(ten: str) -> bool:
    return "giường" in str(ten or "").lower()


# ---------------------------------------------------------------------------
# Rule
# ---------------------------------------------------------------------------
class KiemTraNgayGiuong(GiamDinhBase):
    """
    G1: Số dịch vụ giường ≠ (NGAY_RA - NGAY_VAO + 1)
    G2: Cùng MA_LK có > 1 dịch vụ giường trong cùng ngày
    Chỉ áp dụng MA_LOAI_KCB = 03.
    """

    def check(self, all_objects: dict) -> list[GiamDinhError]:
        errors: list[GiamDinhError] = []

        # XML1: chỉ lấy BN nội trú MA_LOAI_KCB=03
        xml1_map: dict[str, dict] = {}
        for row_excel, rec in self._get_rows(all_objects, "XML1"):
            ma_lk = (getattr(rec, "MA_LK",       None) or "").strip()
            loai  = (getattr(rec, "MA_LOAI_KCB",  None) or "").strip()
            if not ma_lk or loai != MA_LOAI_KCB_NOI_TRU:
                continue
            xml1_map[ma_lk] = {
                "row_excel": row_excel,
                "ngay_vao":  (getattr(rec, "NGAY_VAO", None) or "").strip(),
                "ngay_ra":   (getattr(rec, "NGAY_RA",  None) or "").strip(),
                "ho_ten":    (getattr(rec, "HO_TEN",   None) or "").strip(),
                "ma_bn":     (getattr(rec, "MA_BN",    None) or "").strip(),
            }

        if not xml1_map:
            return errors

        # XML3: thu thập dịch vụ giường
        # giuong_rows[MA_LK] = list[(row_excel, ngay_str, ten_dich_vu)]
        giuong_rows: dict[str, list[tuple]] = defaultdict(list)
        for row_excel, rec in self._get_rows(all_objects, "XML3"):
            ma_lk       = (getattr(rec, "MA_LK",       None) or "").strip()
            ten_dich_vu  = (getattr(rec, "TEN_DICH_VU", None) or "").strip()
            ngay_th_yl   = (getattr(rec, "NGAY_TH_YL", None) or "").strip()
            if not ma_lk or ma_lk not in xml1_map:
                continue
            if not _is_giuong(ten_dich_vu):
                continue
            ngay_str = str(ngay_th_yl)[:8] if ngay_th_yl else ""
            giuong_rows[ma_lk].append((row_excel, ngay_str, ten_dich_vu))

        # Kiểm tra
        for ma_lk, info in xml1_map.items():
            ngay_vao = info["ngay_vao"]
            ngay_ra  = info["ngay_ra"]
            so_ngay  = _so_ngay(ngay_vao, ngay_ra)
            rows_g   = giuong_rows.get(ma_lk, [])
            dem      = len(rows_g)

            # G1
            if so_ngay is not None and dem != so_ngay:
                errors.append(GiamDinhError(
                    sheet="XML1",
                    ma_lk=ma_lk,
                    row_excel=info["row_excel"],
                    ma_ly_do="WARN.GG01",
                    mo_ta=(
                        f"Số dịch vụ giường ({dem}) "
                        f"{'ít hơn' if dem < so_ngay else 'nhiều hơn'} "
                        f"số ngày điều trị ({so_ngay}). "
                        f"NGAY_VAO={ngay_vao} → NGAY_RA={ngay_ra} "
                        f"(= {so_ngay} ngày)."
                    ),
                    can_cu=CAN_CU,
                ))
            elif so_ngay is None and dem == 0:
                errors.append(GiamDinhError(
                    sheet="XML1",
                    ma_lk=ma_lk,
                    row_excel=info["row_excel"],
                    ma_ly_do="WARN.GG01",
                    mo_ta=(
                        "Không tìm thấy dịch vụ giường nào trong XML3. "
                        f"NGAY_VAO={ngay_vao} → NGAY_RA={ngay_ra}."
                    ),
                    can_cu=CAN_CU,
                ))

            # G2
            ngay_groups: dict[str, list[tuple]] = defaultdict(list)
            for row_excel, ngay_str, ten in rows_g:
                ngay_groups[ngay_str].append((row_excel, ten))

            for ngay_str, rows_ngay in ngay_groups.items():
                if len(rows_ngay) > 1:
                    ten_list = [t for _, t in rows_ngay]
                    for row_excel, ten in rows_ngay:
                        errors.append(GiamDinhError(
                            sheet="XML3",
                            ma_lk=ma_lk,
                            row_excel=row_excel,
                            ma_ly_do="WARN.GG02",
                            mo_ta=(
                                f"Ngày {ngay_str}: {len(rows_ngay)} dịch vụ giường "
                                f"cho cùng BN — mỗi ngày chỉ được 1 dịch vụ giường. "
                                f"Các dịch vụ: {'; '.join(ten_list)}."
                            ),
                            can_cu=CAN_CU,
                            ma_dich_vu=ten,
                        ))

        return errors

    # -----------------------------------------------------------------------
    # Export Excel riêng
    # -----------------------------------------------------------------------
    @staticmethod
    def write_excel(excel_file: str, errors: list[GiamDinhError]) -> str:
        """
        Xuất file Excel kết quả kiểm tra ngày giường.
        Tên: <ten_goc>[<timestamp>]_ngaygiuong.xlsx

        Sheet TONG_HOP  – thống kê số lỗi theo mã cảnh báo
        Sheet CHI_TIET  – toàn bộ lỗi
        Sheet WARN_GG01 – thiếu/thừa dịch vụ giường (lỗi XML1)
        Sheet WARN_GG02 – trùng dịch vụ giường cùng ngày (lỗi XML3)
        """
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from collections import Counter

        # Style
        FILL_HDR   = PatternFill("solid", start_color="1F3864")
        FILL_G1    = PatternFill("solid", start_color="FCE4D6")  # cam – G1
        FILL_G2    = PatternFill("solid", start_color="FFF2CC")  # vàng – G2
        FILL_ODD   = PatternFill("solid", start_color="EBF3FB")
        FILL_EVEN  = PatternFill("solid", start_color="FFFFFF")
        FILL_OK    = PatternFill("solid", start_color="E2EFDA")
        FILL_SUM   = PatternFill("solid", start_color="C00000")

        FONT_HDR   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        FONT_DATA  = Font(name="Arial", size=10)
        FONT_RED   = Font(name="Arial", size=10, color="C00000", bold=True)
        FONT_WHITE = Font(name="Arial", size=10, color="FFFFFF", bold=True)

        ALIGN_CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ALIGN_L    = Alignment(vertical="center", wrap_text=True)
        ALIGN_TOP  = Alignment(vertical="top", wrap_text=True)

        _THIN = Side(style="thin", color="AAAAAA")
        BDR   = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

        def hdr(ws, r, c, v, fill=FILL_HDR):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = fill; cell.font = FONT_HDR
            cell.alignment = ALIGN_CTR; cell.border = BDR
            return cell

        def dat(ws, r, c, v, fill=FILL_EVEN, font=FONT_DATA, align=ALIGN_L):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = fill; cell.font = font
            cell.alignment = align; cell.border = BDR
            return cell

        def auto_w(ws, headers, start=2):
            for ci, h in enumerate(headers, 1):
                mx = len(str(h))
                for ri in range(start, ws.max_row + 1):
                    mx = max(mx, min(len(str(ws.cell(ri, ci).value or "").split("\n")[0]), 80))
                ws.column_dimensions[get_column_letter(ci)].width = mx + 3

        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)

        now_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        cnt = Counter(e.ma_ly_do for e in errors)

        # ===================================================================
        # Sheet TONG_HOP
        # ===================================================================
        ws = wb.create_sheet("TONG_HOP")
        ws.merge_cells("A1:D1")
        tc = ws.cell(1, 1, f"KIỂM TRA NGÀY GIƯỜNG NỘI TRÚ (MA_LOAI_KCB=03)  –  {now_str}")
        tc.fill = FILL_HDR; tc.font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
        tc.alignment = ALIGN_CTR; ws.row_dimensions[1].height = 28

        TH_HDRS = ["MÃ CẢNH BÁO", "Ý NGHĨA", "SỐ TRƯỜNG HỢP"]
        for ci, h in enumerate(TH_HDRS, 1):
            hdr(ws, 2, ci, h)

        MA_MO_TA = {
            "WARN.GG01": "Số dịch vụ giường ≠ số ngày điều trị (NGAY_RA − NGAY_VAO + 1)",
            "WARN.GG02": "Cùng BN, cùng ngày có > 1 dịch vụ giường",
        }
        FILL_MA = {"WARN.GG01": FILL_G1, "WARN.GG02": FILL_G2}

        ri = 3
        for ma in ["WARN.GG01", "WARN.GG02"]:
            so = cnt.get(ma, 0)
            fill = FILL_MA.get(ma, FILL_EVEN)
            dat(ws, ri, 1, ma,             fill=fill, align=ALIGN_CTR)
            dat(ws, ri, 2, MA_MO_TA[ma],   fill=fill)
            dat(ws, ri, 3, so,             fill=fill, font=FONT_RED if so else FONT_DATA, align=ALIGN_CTR)
            ri += 1

        dat(ws, ri, 1, "TỔNG",    fill=FILL_SUM, font=FONT_WHITE, align=ALIGN_CTR)
        dat(ws, ri, 2, "",        fill=FILL_SUM)
        dat(ws, ri, 3, len(errors), fill=FILL_SUM, font=FONT_WHITE, align=ALIGN_CTR)

        if not errors:
            ws.merge_cells(f"A{ri+2}:D{ri+2}")
            ok = ws.cell(ri + 2, 1, "✓ Không phát hiện lỗi ngày giường.")
            ok.fill = FILL_OK
            ok.font = Font(name="Arial", size=11, bold=True, color="375623")
            ok.alignment = ALIGN_CTR

        auto_w(ws, TH_HDRS, start=3); ws.freeze_panes = "A3"

        # ===================================================================
        # Sheet CHI_TIET
        # ===================================================================
        ws_ct = wb.create_sheet("CHI_TIET")
        CT_HDRS = ["STT", "MÃ CẢNH BÁO", "SHEET", "ROW_EXCEL",
                   "MA_LK", "MA_DICH_VU / TEN_DV", "MÔ TẢ"]
        for ci, h in enumerate(CT_HDRS, 1):
            hdr(ws_ct, 1, ci, h)

        for idx, err in enumerate(errors, 1):
            ri = idx + 1
            fill = FILL_MA.get(err.ma_ly_do, FILL_ODD if idx % 2 else FILL_EVEN)
            dat(ws_ct, ri, 1, idx,              fill=fill, align=ALIGN_CTR)
            dat(ws_ct, ri, 2, err.ma_ly_do,     fill=fill, font=FONT_RED, align=ALIGN_CTR)
            dat(ws_ct, ri, 3, err.sheet,         fill=fill, align=ALIGN_CTR)
            dat(ws_ct, ri, 4, err.row_excel,     fill=fill, align=ALIGN_CTR)
            dat(ws_ct, ri, 5, err.ma_lk,         fill=fill)
            dat(ws_ct, ri, 6, err.ma_dich_vu,    fill=fill, align=ALIGN_TOP)
            dat(ws_ct, ri, 7, err.mo_ta,         fill=fill, align=ALIGN_TOP)
            ws_ct.row_dimensions[ri].height = max(25, len(err.mo_ta) // 60 * 14 + 14)

        auto_w(ws_ct, CT_HDRS); ws_ct.freeze_panes = "A2"; ws_ct.row_dimensions[1].height = 28

        # ===================================================================
        # Sheet WARN_GG01 – lỗi thiếu/thừa giường (từ XML1)
        # ===================================================================
        g1_errors = [e for e in errors if e.ma_ly_do == "WARN.GG01"]
        ws_g1 = wb.create_sheet("WARN_GG01_SoNgay")
        G1_HDRS = ["STT", "ROW_EXCEL", "MA_LK", "MÔ TẢ"]
        for ci, h in enumerate(G1_HDRS, 1):
            hdr(ws_g1, 1, ci, h, fill=PatternFill("solid", start_color="C55A11"))

        for idx, err in enumerate(g1_errors, 1):
            ri = idx + 1
            fill = FILL_G1 if idx % 2 else FILL_EVEN
            dat(ws_g1, ri, 1, idx,          fill=fill, align=ALIGN_CTR)
            dat(ws_g1, ri, 2, err.row_excel, fill=fill, align=ALIGN_CTR)
            dat(ws_g1, ri, 3, err.ma_lk,    fill=fill)
            dat(ws_g1, ri, 4, err.mo_ta,    fill=fill, align=ALIGN_TOP)
            ws_g1.row_dimensions[ri].height = max(25, len(err.mo_ta) // 60 * 14 + 14)

        if not g1_errors:
            ws_g1.cell(2, 1, "✓ Không có lỗi").fill = FILL_OK

        auto_w(ws_g1, G1_HDRS); ws_g1.freeze_panes = "A2"; ws_g1.row_dimensions[1].height = 28

        # ===================================================================
        # Sheet WARN_GG02 – lỗi trùng giường cùng ngày (từ XML3)
        # ===================================================================
        g2_errors = [e for e in errors if e.ma_ly_do == "WARN.GG02"]
        ws_g2 = wb.create_sheet("WARN_GG02_TrungNgay")
        G2_HDRS = ["STT", "ROW_EXCEL", "MA_LK", "TEN_DICH_VU", "MÔ TẢ"]
        for ci, h in enumerate(G2_HDRS, 1):
            hdr(ws_g2, 1, ci, h, fill=PatternFill("solid", start_color="7F6000"))

        for idx, err in enumerate(g2_errors, 1):
            ri = idx + 1
            fill = FILL_G2 if idx % 2 else FILL_EVEN
            dat(ws_g2, ri, 1, idx,           fill=fill, align=ALIGN_CTR)
            dat(ws_g2, ri, 2, err.row_excel,  fill=fill, align=ALIGN_CTR)
            dat(ws_g2, ri, 3, err.ma_lk,     fill=fill)
            dat(ws_g2, ri, 4, err.ma_dich_vu, fill=fill, align=ALIGN_TOP)
            dat(ws_g2, ri, 5, err.mo_ta,     fill=fill, align=ALIGN_TOP)
            ws_g2.row_dimensions[ri].height = max(25, len(err.mo_ta) // 60 * 14 + 14)

        if not g2_errors:
            ws_g2.cell(2, 1, "✓ Không có lỗi").fill = FILL_OK

        auto_w(ws_g2, G2_HDRS); ws_g2.freeze_panes = "A2"; ws_g2.row_dimensions[1].height = 28

        # ===================================================================
        # Lưu
        # ===================================================================
        folder = os.path.dirname(os.path.abspath(excel_file))
        base   = os.path.splitext(os.path.basename(excel_file))[0]
        ts     = datetime.now().strftime("%Y%m%d_%H%M%S")
        out    = os.path.join(folder, f"{base}[{ts}]_ngaygiuong.xlsx")
        wb.save(out)
        return out