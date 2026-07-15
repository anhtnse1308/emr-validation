"""
Rule – Sàng lọc mã đối tượng KCB (MA_DOITUONG_KCB) tại XML1
Tất cả trường dữ liệu đều từ XML1.

Quy tắc:
  DT1 – MA_DOITUONG_KCB=1.14 → MA_LOAI_KCB phải thuộc {01,02,05,08}   → WARN.DT01
  DT2 – MA_DOITUONG_KCB=1.15 → MA_LOAI_KCB phải thuộc {03,09}         → WARN.DT02
  DT3 – MA_DKBD=74021        → MA_DOITUONG_KCB phải thuộc {1.1, 2}    → WARN.DT03
  DT4 – MA_DKBD≠74021        → MA_DOITUONG_KCB phải thuộc {1.14,1.15,1.3}
         (1.3 yêu cầu có GIAY_CHUYEN_TUYEN)                            → WARN.DT04
  DT5 – MA_DOITUONG_KCB=1.7  → |NGAY_VAO.date - NGAY_SINH.date| ≤ 5  → WARN.DT05
"""

from __future__ import annotations
import os
from datetime import datetime, date

from giamdinh.GiamDinhBase import GiamDinhBase, GiamDinhError

CAN_CU = "Thông tư 30/2020/TT-BYT + QĐ 4210/QĐ-BYT (Mã đối tượng KCB BHYT)"

# ---------------------------------------------------------------------------
# Tập hằng số
# ---------------------------------------------------------------------------
LOAI_KCB_NGOAI_TRU = {"01", "02", "05", "08"}
LOAI_KCB_NOI_TRU   = {"03", "09"}

DKBD_DUNG_TUYEN    = "74021"
DOITUONG_DUNG_TUYEN = {"1.1", "2"}
DOITUONG_TRAI_TUYEN = {"1.14", "1.15", "1.3", "2"}

DOITUONG_SO_SINH    = "1.7"
SO_SINH_MAX_NGAY    = 5


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------
def _parse_date(val: str) -> date | None:
    s = str(val or "").strip()[:8]
    if len(s) == 8 and s.isdigit():
        try:
            return datetime.strptime(s, "%Y%m%d").date()
        except ValueError:
            pass
    return None


def _norm(val) -> str:
    return str(val or "").strip()


# ---------------------------------------------------------------------------
# Rule
# ---------------------------------------------------------------------------
class SangLocDoiTuong(GiamDinhBase):
    """Sàng lọc mã đối tượng KCB (MA_DOITUONG_KCB) theo quy tắc nghiệp vụ."""

    def check(self, all_objects: dict) -> list[GiamDinhError]:
        errors: list[GiamDinhError] = []

        for row_excel, rec in self._get_rows(all_objects, "XML1"):
            ma_lk       = _norm(getattr(rec, "MA_LK",           None))
            doi_tuong   = _norm(getattr(rec, "MA_DOITUONG_KCB",  None))
            loai_kcb    = _norm(getattr(rec, "MA_LOAI_KCB",      None))
            ma_dkbd     = _norm(getattr(rec, "MA_DKBD",          None))
            giay_ct     = _norm(getattr(rec, "GIAY_CHUYEN_TUYEN",None))
            ma_noi_di   = _norm(getattr(rec, "MA_NOI_DI",       None))
            ngay_sinh   = _norm(getattr(rec, "NGAY_SINH",        None))
            ngay_vao    = _norm(getattr(rec, "NGAY_VAO",         None))

            if not ma_lk or not doi_tuong:
                continue

            # ==============================================================
            # DT1 – 1.14 → ngoại trú (01,02,05,08)
            # ==============================================================
            if doi_tuong == "1.14" and loai_kcb not in LOAI_KCB_NGOAI_TRU:
                errors.append(GiamDinhError(
                    sheet="XML1", ma_lk=ma_lk, row_excel=row_excel,
                    ma_ly_do="WARN.DT01",
                    mo_ta=(
                        f"MA_DOITUONG_KCB=1.14 (trái tuyến ngoại trú) "
                        f"nhưng MA_LOAI_KCB='{loai_kcb}' không hợp lệ. "
                        f"Phải thuộc {{01,02,05,08}}."
                    ),
                    can_cu=CAN_CU,
                ))

            # ==============================================================
            # DT2 – 1.15 → nội trú (03,09)
            # ==============================================================
            if doi_tuong == "1.15" and loai_kcb not in LOAI_KCB_NOI_TRU:
                errors.append(GiamDinhError(
                    sheet="XML1", ma_lk=ma_lk, row_excel=row_excel,
                    ma_ly_do="WARN.DT02",
                    mo_ta=(
                        f"MA_DOITUONG_KCB=1.15 (chuyển từ 1.14 trái tuyến sang nội trú) "
                        f"nhưng MA_LOAI_KCB='{loai_kcb}' không hợp lệ. "
                        f"Phải thuộc {{03,09}}."
                    ),
                    can_cu=CAN_CU,
                ))

            # ==============================================================
            # DT3 – MA_DKBD=74021 → đối tượng phải là 1.1 hoặc 2
            # ==============================================================
            if ma_dkbd == DKBD_DUNG_TUYEN and doi_tuong not in DOITUONG_DUNG_TUYEN:
                errors.append(GiamDinhError(
                    sheet="XML1", ma_lk=ma_lk, row_excel=row_excel,
                    ma_ly_do="WARN.DT03",
                    mo_ta=(
                        f"MA_DKBD='{ma_dkbd}' (đúng tuyến CSKCB đăng ký) "
                        f"nhưng MA_DOITUONG_KCB='{doi_tuong}' không hợp lệ. "
                        f"Phải là 1.1 hoặc 2."
                    ),
                    can_cu=CAN_CU,
                ))

            # ==============================================================
            # DT4 – MA_DKBD≠74021 → đối tượng phải là 1.14, 1.15, hoặc 1.3
            #        Nếu 1.3 thì phải có GIAY_CHUYEN_TUYEN
            # ==============================================================
            if ma_dkbd and ma_dkbd != DKBD_DUNG_TUYEN:
                if doi_tuong not in DOITUONG_TRAI_TUYEN:
                    errors.append(GiamDinhError(
                        sheet="XML1", ma_lk=ma_lk, row_excel=row_excel,
                        ma_ly_do="WARN.DT04",
                        mo_ta=(
                            f"MA_DKBD='{ma_dkbd}' (khác nơi đăng ký) "
                            f"nhưng MA_DOITUONG_KCB='{doi_tuong}' không hợp lệ. "
                            f"Phải là 1.14 (ngoại trú), 1.15 (nội trú) "
                            f"hoặc 1.3 (có giấy chuyển tuyến)."
                        ),
                        can_cu=CAN_CU,
                    ))
                    """
                elif doi_tuong == "1.3" and not giay_ct:
                    errors.append(GiamDinhError(
                        sheet="XML1", ma_lk=ma_lk, row_excel=row_excel,
                        ma_ly_do="WARN.DT04",
                        mo_ta=(
                            f"MA_DOITUONG_KCB=1.3 (có giấy chuyển tuyến) "
                            f"nhưng GIAY_CHUYEN_TUYEN để trống. "
                            f"Phải ghi số giấy chuyển tuyến."
                        ),
                        can_cu=CAN_CU,
                    ))
                    """
                elif doi_tuong == "1.3" and not ma_noi_di:
                    errors.append(GiamDinhError(
                        sheet="XML1", ma_lk=ma_lk, row_excel=row_excel,
                        ma_ly_do="WARN.DT04",
                        mo_ta=(
                            f"MA_DOITUONG_KCB=1.3 (có giấy chuyển tuyến) "
                            f"nhưng MA_NOI_DI để trống. "
                            f"Phải ghi MA_NOI_DI tức nơi chuyển tuyến."
                        ),
                        can_cu=CAN_CU,
                    ))

            # ==============================================================
            # DT5 – 1.7 (sơ sinh) → NGAY_VAO - NGAY_SINH ≤ 5 ngày
            # ==============================================================
            if doi_tuong == DOITUONG_SO_SINH:
                d_sinh = _parse_date(ngay_sinh)
                d_vao  = _parse_date(ngay_vao)
                if d_sinh and d_vao:
                    delta = abs((d_vao - d_sinh).days)
                    if delta > SO_SINH_MAX_NGAY:
                        errors.append(GiamDinhError(
                            sheet="XML1", ma_lk=ma_lk, row_excel=row_excel,
                            ma_ly_do="WARN.DT05",
                            mo_ta=(
                                f"MA_DOITUONG_KCB=1.7 (trẻ sơ sinh) "
                                f"nhưng NGAY_VAO ({ngay_vao[:8]}) cách "
                                f"NGAY_SINH ({ngay_sinh[:8]}) {delta} ngày "
                                f"(> {SO_SINH_MAX_NGAY} ngày cho phép)."
                            ),
                            can_cu=CAN_CU,
                        ))
                elif doi_tuong == DOITUONG_SO_SINH and (not d_sinh or not d_vao):
                    errors.append(GiamDinhError(
                        sheet="XML1", ma_lk=ma_lk, row_excel=row_excel,
                        ma_ly_do="WARN.DT05",
                        mo_ta=(
                            f"MA_DOITUONG_KCB=1.7 (trẻ sơ sinh) "
                            f"nhưng không thể xác định ngày: "
                            f"NGAY_SINH='{ngay_sinh}', NGAY_VAO='{ngay_vao}'."
                        ),
                        can_cu=CAN_CU,
                    ))

        return errors

    # -----------------------------------------------------------------------
    # Export Excel riêng
    # -----------------------------------------------------------------------
    @staticmethod
    def write_excel(excel_file: str, errors: list[GiamDinhError]) -> str:
        """
        Xuất kết quả sàng lọc đối tượng KCB.
        Tên: <ten_goc>[<timestamp>]_sanglocdoituong.xlsx

        Sheet TONG_HOP  – thống kê theo mã cảnh báo
        Sheet CHI_TIET  – toàn bộ lỗi
        Sheet per mã    – WARN_DT01..DT05
        """
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from collections import Counter

        FILLS = {
            "HDR":      PatternFill("solid", start_color="1F3864"),
            "DT01":     PatternFill("solid", start_color="FCE4D6"),
            "DT02":     PatternFill("solid", start_color="FFF2CC"),
            "DT03":     PatternFill("solid", start_color="DDEBF7"),
            "DT04":     PatternFill("solid", start_color="EAF1DD"),
            "DT05":     PatternFill("solid", start_color="F4CCFF"),
            "OK":       PatternFill("solid", start_color="E2EFDA"),
            "SUM":      PatternFill("solid", start_color="C00000"),
            "EVEN":     PatternFill("solid", start_color="FFFFFF"),
        }

        HDR_COLORS = {
            "WARN.DT01": "C55A11",
            "WARN.DT02": "7F6000",
            "WARN.DT03": "1F497D",
            "WARN.DT04": "375623",
            "WARN.DT05": "5B2C6F",
        }

        FONT_HDR   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        FONT_DATA  = Font(name="Arial", size=10)
        FONT_RED   = Font(name="Arial", size=10, color="C00000", bold=True)
        FONT_WHITE = Font(name="Arial", size=10, color="FFFFFF", bold=True)

        ALIGN_CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ALIGN_L    = Alignment(vertical="center", wrap_text=True)
        ALIGN_TOP  = Alignment(vertical="top", wrap_text=True)

        _THIN = Side(style="thin", color="AAAAAA")
        BDR   = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

        def hdr(ws, r, c, v, fill=None):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = fill or FILLS["HDR"]
            cell.font = FONT_HDR
            cell.alignment = ALIGN_CTR
            cell.border = BDR
            return cell

        def dat(ws, r, c, v, fill=FILLS["EVEN"], font=FONT_DATA, align=ALIGN_L):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = fill; cell.font = font
            cell.alignment = align; cell.border = BDR
            return cell

        def auto_w(ws, headers, start=2):
            for ci, h in enumerate(headers, 1):
                mx = len(str(h))
                for ri in range(start, ws.max_row + 1):
                    mx = max(mx, min(len(str(ws.cell(ri, ci).value or "").split("\n")[0]), 80))
                ws.column_dimensions[get_column_letter(ci)].width = mx + 3

        MA_MO_TA = {
            "WARN.DT01": "MA_DOITUONG_KCB=1.14 nhưng MA_LOAI_KCB không phải ngoại trú (01,02,05,08)",
            "WARN.DT02": "MA_DOITUONG_KCB=1.15 nhưng MA_LOAI_KCB không phải nội trú (03,09)",
            "WARN.DT03": "MA_DKBD=74021 (đúng tuyến) nhưng MA_DOITUONG_KCB không phải 1.1 hoặc 2",
            "WARN.DT04": "MA_DKBD≠74021 (trái tuyến) nhưng MA_DOITUONG_KCB sai hoặc thiếu giấy chuyển",
            "WARN.DT05": "MA_DOITUONG_KCB=1.7 (sơ sinh) nhưng NGAY_VAO cách NGAY_SINH > 5 ngày",
        }

        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)

        now_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        cnt = Counter(e.ma_ly_do for e in errors)

        # TONG_HOP
        ws = wb.create_sheet("TONG_HOP")
        ws.merge_cells("A1:D1")
        tc = ws.cell(1, 1, f"SÀNG LỌC MÃ ĐỐI TƯỢNG KCB (MA_DOITUONG_KCB)  –  {now_str}")
        tc.fill = FILLS["HDR"]
        tc.font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
        tc.alignment = ALIGN_CTR; ws.row_dimensions[1].height = 28

        TH_HDRS = ["MÃ CẢNH BÁO", "QUY TẮC VI PHẠM", "SỐ TRƯỜNG HỢP"]
        for ci, h in enumerate(TH_HDRS, 1):
            hdr(ws, 2, ci, h)

        ri = 3
        for ma, mo_ta in MA_MO_TA.items():
            so   = cnt.get(ma, 0)
            key  = ma.replace("WARN.", "")
            fill = FILLS.get(key, FILLS["EVEN"])
            dat(ws, ri, 1, ma,    fill=fill, align=ALIGN_CTR)
            dat(ws, ri, 2, mo_ta, fill=fill)
            dat(ws, ri, 3, so,    fill=fill,
                font=FONT_RED if so else FONT_DATA, align=ALIGN_CTR)
            ri += 1

        dat(ws, ri, 1, "TỔNG",      fill=FILLS["SUM"], font=FONT_WHITE, align=ALIGN_CTR)
        dat(ws, ri, 2, "",          fill=FILLS["SUM"])
        dat(ws, ri, 3, len(errors), fill=FILLS["SUM"], font=FONT_WHITE, align=ALIGN_CTR)

        if not errors:
            ws.merge_cells(f"A{ri+2}:D{ri+2}")
            ok = ws.cell(ri + 2, 1, "✓ Không phát hiện vi phạm mã đối tượng KCB.")
            ok.fill = FILLS["OK"]
            ok.font = Font(name="Arial", size=11, bold=True, color="375623")
            ok.alignment = ALIGN_CTR

        auto_w(ws, TH_HDRS, start=3); ws.freeze_panes = "A3"

        # CHI_TIET
        ws_ct = wb.create_sheet("CHI_TIET")
        CT_HDRS = ["STT", "MÃ CẢNH BÁO", "ROW_EXCEL", "MA_LK", "MÔ TẢ"]
        for ci, h in enumerate(CT_HDRS, 1):
            hdr(ws_ct, 1, ci, h)

        for idx, err in enumerate(errors, 1):
            ri = idx + 1
            key  = err.ma_ly_do.replace("WARN.", "")
            fill = FILLS.get(key, FILLS["EVEN"] if idx % 2 == 0 else PatternFill("solid", start_color="EBF3FB"))
            dat(ws_ct, ri, 1, idx,           fill=fill, align=ALIGN_CTR)
            dat(ws_ct, ri, 2, err.ma_ly_do,  fill=fill, font=FONT_RED, align=ALIGN_CTR)
            dat(ws_ct, ri, 3, err.row_excel,  fill=fill, align=ALIGN_CTR)
            dat(ws_ct, ri, 4, err.ma_lk,     fill=fill)
            dat(ws_ct, ri, 5, err.mo_ta,     fill=fill, align=ALIGN_TOP)
            ws_ct.row_dimensions[ri].height = max(25, len(err.mo_ta) // 60 * 14 + 14)

        auto_w(ws_ct, CT_HDRS); ws_ct.freeze_panes = "A2"; ws_ct.row_dimensions[1].height = 28

        # Sheet per mã cảnh báo
        groups: dict[str, list[GiamDinhError]] = {}
        for e in errors:
            groups.setdefault(e.ma_ly_do, []).append(e)

        for ma_ly_do, grp in sorted(groups.items()):
            sname = ma_ly_do.replace("WARN.", "WARN_").replace(".", "_")[:31]
            ws_g  = wb.create_sheet(sname)
            G_HDRS = ["STT", "ROW_EXCEL", "MA_LK", "MÔ TẢ"]
            hdr_color = HDR_COLORS.get(ma_ly_do, "1F3864")
            for ci, h in enumerate(G_HDRS, 1):
                hdr(ws_g, 1, ci, h,
                    fill=PatternFill("solid", start_color=hdr_color))

            key  = ma_ly_do.replace("WARN.", "")
            fill_odd = FILLS.get(key, PatternFill("solid", start_color="EBF3FB"))

            for idx, err in enumerate(grp, 1):
                ri = idx + 1
                fill = fill_odd if idx % 2 else FILLS["EVEN"]
                dat(ws_g, ri, 1, idx,          fill=fill, align=ALIGN_CTR)
                dat(ws_g, ri, 2, err.row_excel, fill=fill, align=ALIGN_CTR)
                dat(ws_g, ri, 3, err.ma_lk,    fill=fill)
                dat(ws_g, ri, 4, err.mo_ta,    fill=fill, align=ALIGN_TOP)
                ws_g.row_dimensions[ri].height = max(25, len(err.mo_ta) // 60 * 14 + 14)

            if not grp:
                ws_g.cell(2, 1, "✓ Không có lỗi").fill = FILLS["OK"]

            auto_w(ws_g, G_HDRS); ws_g.freeze_panes = "A2"; ws_g.row_dimensions[1].height = 28

        # Lưu
        folder = os.path.dirname(os.path.abspath(excel_file))
        base   = os.path.splitext(os.path.basename(excel_file))[0]
        ts     = datetime.now().strftime("%Y%m%d_%H%M%S")
        out    = os.path.join(folder, f"{base}[{ts}]_sanglocdoituong.xlsx")
        wb.save(out)
        return out