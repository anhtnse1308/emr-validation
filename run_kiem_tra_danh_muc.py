"""
Luồng kiểm tra thuốc XML2 theo danh mục trúng thầu.
Chạy độc lập, không phụ thuộc vào GiamDinhService chính.

Dùng:
    python run_kiem_tra_danh_muc.py

Sẽ hỏi:
    1. Đường dẫn file Excel BHYT (chứa sheet XML2)
    2. Đường dẫn file danh mục thuốc (FileDanhMucThuoc.xlsx)

Output:
    <ten_file_bhyt>[<timestamp>]_danhmuc.xlsx
"""

import os
import sys
from datetime import datetime
from collections import Counter

# Cho phép chạy từ bất kỳ thư mục nào
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from services.excel_service import ExcelService
from giamdinh.DanhMucThuocLoader import DanhMucThuocLoader
from giamdinh.rules.KiemTraThuocDanhMuc import KiemTraThuocDanhMuc
from giamdinh.GiamDinhBase import GiamDinhError


# ===========================================================================
# Input
# ===========================================================================
def _prompt_path(label: str) -> str:
    while True:
        try:
            p = input(f"{label}: ").strip().strip('"')
        except KeyboardInterrupt:
            print("\nĐã hủy.")
            sys.exit(0)
        if os.path.exists(p):
            return p
        choice = input("❌ File không tồn tại! Nhập lại? (Y/N): ").strip().lower()
        if choice != "y":
            print("Đã hủy.")
            sys.exit(0)


# ===========================================================================
# Excel export
# ===========================================================================
def write_excel_danhmuc(
    excel_file: str,
    errors: list[GiamDinhError],
    danh_muc_path: str,
) -> str:
    """Xuất kết quả kiểm tra danh mục ra Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    FILL_HDR   = PatternFill("solid", start_color="1F3864")
    FILL_ODD   = PatternFill("solid", start_color="FFF2CC")
    FILL_EVEN  = PatternFill("solid", start_color="FFFFFF")
    FILL_WARN  = PatternFill("solid", start_color="FCE4D6")
    FILL_OK    = PatternFill("solid", start_color="E2EFDA")
    FILL_SUM   = PatternFill("solid", start_color="C00000")

    FONT_HDR   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    FONT_DATA  = Font(name="Arial", size=10)
    FONT_RED   = Font(name="Arial", size=10, color="C00000", bold=True)
    FONT_WARN  = Font(name="Arial", size=10, color="833C00")

    ALIGN_CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ALIGN_L    = Alignment(vertical="center", wrap_text=True)
    ALIGN_TOP  = Alignment(vertical="top", wrap_text=True)

    _THIN = Side(style="thin", color="AAAAAA")
    BDR   = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

    def hdr(ws, r, c, v):
        cell = ws.cell(row=r, column=c, value=v)
        cell.fill = FILL_HDR; cell.font = FONT_HDR
        cell.alignment = ALIGN_CTR; cell.border = BDR
        return cell

    def cell(ws, r, c, v, fill=FILL_EVEN, font=FONT_DATA, align=ALIGN_L):
        cell = ws.cell(row=r, column=c, value=v)
        cell.fill = fill; cell.font = font
        cell.alignment = align; cell.border = BDR
        return cell

    def auto_width(ws, headers, data_start_row=2):
        for ci, h in enumerate(headers, 1):
            max_w = len(str(h))
            for ri in range(data_start_row, ws.max_row + 1):
                v = ws.cell(row=ri, column=ci).value or ""
                max_w = max(max_w, min(len(str(v).split("\n")[0]), 80))
            ws.column_dimensions[get_column_letter(ci)].width = max_w + 3

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    now_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    # -----------------------------------------------------------------------
    # Sheet 1: TONG_HOP
    # -----------------------------------------------------------------------
    ws_th = wb.create_sheet("TONG_HOP")
    ws_th.merge_cells("A1:E1")
    tc = ws_th.cell(row=1, column=1,
        value=f"KIỂM TRA THUỐC THEO DANH MỤC TRÚNG THẦU  –  {now_str}")
    tc.fill = FILL_HDR; tc.font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
    tc.alignment = ALIGN_CTR
    ws_th.row_dimensions[1].height = 28

    ws_th.cell(row=2, column=1, value=f"Danh mục: {danh_muc_path}").font = FONT_DATA
    ws_th.merge_cells("A2:E2")

    TH_HDRS = ["MÃ CẢNH BÁO", "Ý NGHĨA", "SỐ DÒNG"]
    for ci, h in enumerate(TH_HDRS, 1):
        hdr(ws_th, 3, ci, h)

    MA_MOT_TA = {
        "WARN.DM01": "MA_THUOC không có trong danh mục",
        "WARN.DM02": "SO_DANG_KY không khớp với MA_THUOC trong danh mục",
        "WARN.DM03": "DON_GIA vượt DON_GIA_BH trong danh mục",
        "WARN.DM04": "DUONG_DUNG không khớp với danh mục",
        "WARN.DM05": "HAM_LUONG không khớp với danh mục",
        "WARN.DM06": "Thuốc hết hiệu lực tại thời điểm y lệnh",
    }

    cnt = Counter(e.ma_ly_do for e in errors)
    ri = 4
    for ma, so_luong in sorted(cnt.items()):
        fill = FILL_WARN if so_luong else FILL_OK
        cell(ws_th, ri, 1, ma,      fill=fill, font=FONT_WARN, align=ALIGN_CTR)
        cell(ws_th, ri, 2, MA_MOT_TA.get(ma, ""), fill=fill)
        cell(ws_th, ri, 3, so_luong, fill=fill, font=FONT_RED, align=ALIGN_CTR)
        ri += 1

    cell(ws_th, ri, 1, "TỔNG", fill=FILL_SUM,
         font=Font(bold=True, color="FFFFFF", name="Arial"), align=ALIGN_CTR)
    cell(ws_th, ri, 2, "", fill=FILL_SUM)
    cell(ws_th, ri, 3, len(errors), fill=FILL_SUM,
         font=Font(bold=True, color="FFFFFF", name="Arial"), align=ALIGN_CTR)

    if not errors:
        msg = ws_th.cell(row=ri + 2, column=1,
            value="✓ Không phát hiện lỗi so với danh mục thuốc trúng thầu.")
        msg.fill = FILL_OK
        msg.font = Font(name="Arial", size=11, bold=True, color="375623")
        msg.alignment = ALIGN_CTR
        ws_th.merge_cells(f"A{ri+2}:E{ri+2}")

    auto_width(ws_th, TH_HDRS, data_start_row=3)
    ws_th.freeze_panes = "A4"

    # -----------------------------------------------------------------------
    # Sheet 2: CHI_TIET
    # -----------------------------------------------------------------------
    ws_ct = wb.create_sheet("CHI_TIET")
    CT_HDRS = ["STT", "SHEET", "ROW_EXCEL", "MA_LK", "MA_THUOC",
               "MA_LY_DO", "MO_TA", "CAN_CU"]
    for ci, h in enumerate(CT_HDRS, 1):
        hdr(ws_ct, 1, ci, h)

    for idx, err in enumerate(errors, 1):
        ri = idx + 1
        fill = FILL_ODD if idx % 2 else FILL_EVEN
        cell(ws_ct, ri, 1, idx,           fill=fill, align=ALIGN_CTR)
        cell(ws_ct, ri, 2, err.sheet,     fill=fill, align=ALIGN_CTR)
        cell(ws_ct, ri, 3, err.row_excel, fill=fill, align=ALIGN_CTR)
        cell(ws_ct, ri, 4, err.ma_lk,     fill=fill)
        cell(ws_ct, ri, 5, err.ma_dich_vu, fill=fill)
        cell(ws_ct, ri, 6, err.ma_ly_do,  fill=fill,
             font=FONT_RED, align=ALIGN_CTR)
        cell(ws_ct, ri, 7, err.mo_ta,     fill=fill, align=ALIGN_TOP)
        cell(ws_ct, ri, 8, err.can_cu,    fill=fill, align=ALIGN_TOP)
        ws_ct.row_dimensions[ri].height = max(
            25, len(err.mo_ta) // 60 * 14 + 14
        )

    auto_width(ws_ct, CT_HDRS)
    ws_ct.freeze_panes = "A2"
    ws_ct.row_dimensions[1].height = 28

    # -----------------------------------------------------------------------
    # Sheet per mã cảnh báo
    # -----------------------------------------------------------------------
    groups: dict[str, list[GiamDinhError]] = {}
    for e in errors:
        groups.setdefault(e.ma_ly_do, []).append(e)

    for ma_ly_do, grp in sorted(groups.items()):
        sname = ma_ly_do.replace(".", "_")[:31]
        ws_g = wb.create_sheet(sname)
        G_HDRS = ["STT", "ROW_EXCEL", "MA_LK", "MA_THUOC", "MO_TA"]
        for ci, h in enumerate(G_HDRS, 1):
            hdr(ws_g, 1, ci, h)
        for idx, err in enumerate(grp, 1):
            ri = idx + 1
            fill = FILL_ODD if idx % 2 else FILL_EVEN
            cell(ws_g, ri, 1, idx,           fill=fill, align=ALIGN_CTR)
            cell(ws_g, ri, 2, err.row_excel, fill=fill, align=ALIGN_CTR)
            cell(ws_g, ri, 3, err.ma_lk,     fill=fill)
            cell(ws_g, ri, 4, err.ma_dich_vu, fill=fill)
            cell(ws_g, ri, 5, err.mo_ta,     fill=fill, align=ALIGN_TOP)
            ws_g.row_dimensions[ri].height = max(
                25, len(err.mo_ta) // 60 * 14 + 14
            )
        auto_width(ws_g, G_HDRS)
        ws_g.freeze_panes = "A2"
        ws_g.row_dimensions[1].height = 28

    # -----------------------------------------------------------------------
    # Lưu
    # -----------------------------------------------------------------------
    folder  = os.path.dirname(os.path.abspath(excel_file))
    base    = os.path.splitext(os.path.basename(excel_file))[0]
    ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
    out     = os.path.join(folder, f"{base}[{ts}]_danhmuc.xlsx")
    wb.save(out)
    return out


# ===========================================================================
# Main
# ===========================================================================
if __name__ == "__main__":
    print("=" * 60)
    print("  KIỂM TRA THUỐC XML2 THEO DANH MỤC TRÚNG THẦU")
    print("=" * 60)

    excel_file   = _prompt_path("Nhập file Excel BHYT")
    danh_muc_file = _prompt_path("Nhập file danh mục thuốc (FileDanhMucThuoc.xlsx)")

    # Load danh mục
    print(f"\nĐang nạp danh mục thuốc: {danh_muc_file} ...")
    loader = DanhMucThuocLoader(danh_muc_file)
    print(f"  → {loader}")

    # Đọc Excel BHYT
    print(f"\nĐang đọc dữ liệu BHYT: {excel_file} ...")
    all_objects, _ = ExcelService.read_excel(excel_file)
    n_xml2 = len(all_objects.get("XML2", []))
    print(f"  → XML2: {n_xml2} dòng")

    if n_xml2 == 0:
        print("Không có dữ liệu XML2. Kết thúc.")
        sys.exit(0)

    # Chạy kiểm tra
    print("\nĐang kiểm tra ...")
    rule   = KiemTraThuocDanhMuc(loader)
    errors = rule.check(all_objects)

    # In tóm tắt
    print(f"\n{'=' * 60}")
    if errors:
        cnt = Counter(e.ma_ly_do for e in errors)
        print(f"⚠️  Phát hiện {len(errors)} dòng cần xem xét:")
        MA_MOT_TA = {
            "WARN.DM01": "MA_THUOC không có trong danh mục",
            "WARN.DM02": "SO_DANG_KY không khớp",
            "WARN.DM03": "DON_GIA vượt danh mục",
            "WARN.DM04": "DUONG_DUNG không khớp",
            "WARN.DM05": "HAM_LUONG không khớp",
            "WARN.DM06": "Thuốc hết hiệu lực",
        }
        for ma, so in sorted(cnt.items()):
            print(f"   {ma:15s} ({MA_MOT_TA.get(ma, '')}): {so} dòng")
    else:
        print("✓ Không phát hiện lỗi so với danh mục thuốc trúng thầu.")

    # Xuất Excel
    out = write_excel_danhmuc(excel_file, errors, danh_muc_file)
    print(f"\n✅ Đã xuất Excel kết quả: {out}")