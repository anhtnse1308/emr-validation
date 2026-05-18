from datetime import datetime
from typing import Optional

# ===========================================================================
# EXCEL_STYLE – Cấu hình style chung cho write_excel_summary + write_excel_errors_only
# Chỉnh màu/font/border tại đây, áp dụng cho cả 2 hàm.
# ===========================================================================
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
 
class EXCEL_STYLE:
    # --- Màu nền ---
    FILL_HDR_SUMMARY  = PatternFill("solid", start_color="2F4F8F")  # header summary  – xanh đậm
    FILL_HDR_ERRORS   = PatternFill("solid", start_color="8B0000")  # header errors   – đỏ đậm
    FILL_ROW_ERR      = PatternFill("solid", start_color="FFCCCC")  # dòng lỗi        – đỏ nhạt
    FILL_ROW_OK       = PatternFill("solid", start_color="CCFFCC")  # dòng OK         – xanh nhạt
    FILL_SHEET_CLEAN  = PatternFill("solid", start_color="E8F5E9")  # sheet không lỗi – xanh rất nhạt
 
    # --- Font ---
    FONT_HDR  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    FONT_DATA = Font(name="Arial", size=10)
    FONT_ERR  = Font(name="Arial", size=10, color="CC0000")          # text lỗi – đỏ
    FONT_CLEAN = Font(italic=True, color="2E7D32", name="Arial", size=10)  # sheet sạch – xanh lá
 
    # --- Alignment ---
    ALIGN_CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ALIGN_L   = Alignment(vertical="center", wrap_text=True)
 
    # --- Border ---
    _THIN  = Side(style="thin",   color="AAAAAA")
    BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
 
    # --- Kích thước ---
    COL_WIDTH_MAX  = 60   # độ rộng cột tối đa (ký tự)
    COL_WIDTH_PAD  = 3    # padding thêm vào sau auto-width
    ROW_HDR_HEIGHT = 30   # chiều cao dòng header (pt)

def to_float(value) -> Optional[float]:
    """
    Chuyển string/int → float. Trả None nếu rỗng hoặc không parse được.
    """
    if value is None or str(value).strip() == "":
        return None
    try:
        return float(str(value).strip())
    except (ValueError, TypeError):
        return None


def decimal_places(value) -> int:
    """
    Đếm số chữ số thập phân của chuỗi số (dùng dấu '.').
    VD: '3.500' → 3, '100' → 0
    """
    s = str(value).strip()
    return len(s.split(".")[1]) if "." in s else 0




# ===========================================================================
# Shared parse helpers – dùng chung cho toàn bộ XML model classes
# Import vào BHYTBase để các class con dùng qua self._parse_date12() v.v.
# ===========================================================================

def parse_date12(value) -> Optional[datetime]:
    """
    Parse chuỗi DATE12 (yyyymmddHHMM, 12 ký tự) → datetime.
    Trả None nếu rỗng, sai độ dài, hoặc ngày giờ không hợp lệ.
    """
    if not value or len(str(value).strip()) != 12:
        return None
    try:
        return datetime.strptime(str(value).strip(), "%Y%m%d%H%M")
    except ValueError:
        return None


def parse_date8(value) -> Optional[datetime]:
    """
    Parse chuỗi DATE8 (yyyymmdd, 8 ký tự) → datetime.
    Trả None nếu rỗng, sai độ dài, hoặc ngày không hợp lệ.
    """
    if not value or len(str(value).strip()) != 8:
        return None
    try:
        return datetime.strptime(str(value).strip(), "%Y%m%d")
    except ValueError:
        return None


def to_int(value) -> Optional[int]:
    """
    Chuyển string/float → int. Trả None nếu rỗng hoặc không parse được.
    Dùng qua float() trước để xử lý cả "5.0", "3.500", v.v.
    """
    if value is None or str(value).strip() == "":
        return None
    try:
        return int(float(str(value).strip()))
    except (ValueError, TypeError):
        return None


import os
import json
from datetime import datetime


class FileHelper:

    @staticmethod
    def write_log(excel_file: str, logs: list[str]) -> str:
        """Ghi log dạng text thuần (dùng cho log tổng hợp nhanh)."""
        folder = os.path.dirname(excel_file)
        base_name = os.path.splitext(os.path.basename(excel_file))[0]
        now = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_file = os.path.join(folder, f"{base_name}[{now}].txt")
        with open(log_file, "w", encoding="utf-8") as f:
            for line in logs:
                f.write(line + "\n")
        return log_file

    @staticmethod
    def write_log_json(
        excel_file: str,
        all_objects: dict,
        suffix: str = "errors",
    ) -> str:
        """
        Ghi log lỗi dạng JSON kèm toàn bộ dữ liệu của dòng lỗi.

        Cấu trúc file output:
        {
            "file": "KB-9-5-26.xlsx",
            "generated_at": "2026-05-12T10:30:00",
            "summary": { "total_rows": 120, "error_rows": 5, "ok_rows": 115 },
            "sheets": {
                "XML1": {
                    "total": 50,
                    "errors": [
                        {
                            "row_excel": 3,
                            "errors": ["[MA_LK] Bắt buộc nhập", ...],
                            "data": { "MA_LK": "", "HO_TEN": "Nguyen Van A", ... }
                        },
                        ...
                    ]
                },
                ...
            }
        }

        Chỉ ghi các dòng CÓ LỖI vào mảng "errors";
        dòng hợp lệ chỉ được đếm vào "total".
        """
        folder = os.path.dirname(os.path.abspath(excel_file))
        base_name = os.path.splitext(os.path.basename(excel_file))[0]
        now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_file = os.path.join(folder, f"{base_name}[{now_str}]_{suffix}.json")

        total_rows = 0
        error_rows = 0
        sheets_data: dict = {}

        for sheet_name, objects in all_objects.items():
            sheet_errors = []
            for excel_row, record in enumerate(objects, start=2):
                total_rows += 1
                errs = record.validate()
                if errs:
                    error_rows += 1
                    # Ghi lại toàn bộ giá trị của dòng dữ liệu
                    row_data = {
                        k: v for k, v in record.__dict__.items()
                        if not k.startswith("_")
                    }
                    sheet_errors.append({
                        "row_excel": excel_row,
                        "errors": errs,
                        "data": row_data,
                    })

            sheets_data[sheet_name] = {
                "total": len(objects),
                "error_count": len(sheet_errors),
                "errors": sheet_errors,
            }

        output = {
            "file": os.path.basename(excel_file),
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "summary": {
                "total_rows": total_rows,
                "error_rows": error_rows,
                "ok_rows": total_rows - error_rows,
            },
            "sheets": sheets_data,
        }

        with open(log_file, "w", encoding="utf-8") as f:
            json.dump(output, f, ensure_ascii=False, indent=2)

        return log_file

    @staticmethod
    def _render_summary_lines(all_objects: dict) -> list[str]:
        """
        Nội dung tổng hợp dùng chung cho cả console lẫn file.
        Trả về list[str], mỗi phần tử là một dòng (không có \\n cuối).
        """
        lines = []
        total_rows = sum(len(v) for v in all_objects.values())
        total_errors = 0

        for sheet_name, objects in all_objects.items():
            sheet_err = sum(1 for r in objects if r.validate())
            total_errors += sheet_err

        lines.append(f"Thời gian: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
        lines.append(
            f"Tổng: {total_rows} dòng | "
            f"Lỗi: {total_errors} | "
            f"OK: {total_rows - total_errors}"
        )

        for sheet_name, objects in all_objects.items():
            lines.append(f"\n{'=' * 60}")
            lines.append(f"  SHEET: {sheet_name}  ({len(objects)} dòng)")
            lines.append(f"{'=' * 60}")
            has_error = False
            for i, record in enumerate(objects, start=2):
                errors = record.validate()
                if errors:
                    has_error = True
                    lines.append(f"\n  [LOI] Row Excel {i}:")
                    for e in errors:
                        lines.append(f"       - {e}")
                else:
                    lines.append(f"  [OK]  Row Excel {i}")
            if not has_error:
                lines.append("  -> Khong co loi.")

        return lines

    @staticmethod
    def write_log_summary(excel_file: str, all_objects: dict) -> str:
        """
        Ghi file log tổng hợp dạng text, nội dung y chang in ra console.
        Tên file: <ten_excel>[<timestamp>]_summary.txt
        """
        folder = os.path.dirname(os.path.abspath(excel_file))
        base_name = os.path.splitext(os.path.basename(excel_file))[0]
        now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_file = os.path.join(folder, f"{base_name}[{now_str}]_summary.txt")

        lines = FileHelper._render_summary_lines(all_objects)
        with open(log_file, "w", encoding="utf-8") as f:
            for line in lines:
                f.write(line + "\n")

        return log_file

    @staticmethod
    def write_log_errors_only(excel_file: str, all_objects: dict) -> str:
        """
        Ghi file txt CHỈ chứa các dòng có lỗi (bỏ qua dòng [OK]).

        Cấu trúc:
        ============================================================
          SHEET: XML3  (50 dòng | 3 lỗi)
        ============================================================
          [LOI] Row Excel 4:
               - MA_NHOM: '20' không hợp lệ, phải trong 1–19
               - THANH_TIEN_BV: kỳ vọng 150000 ...
          [LOI] Row Excel 11:
               - ...

        -> Không có lỗi.    ← chỉ in dòng này khi sheet sạch

        Tên file: <ten_excel>[<timestamp>]_errors_only.txt
        """
        folder = os.path.dirname(os.path.abspath(excel_file))
        base_name = os.path.splitext(os.path.basename(excel_file))[0]
        now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_file = os.path.join(
            folder, f"{base_name}[{now_str}]_errors_only.txt"
        )

        total_rows = sum(len(v) for v in all_objects.values())
        total_error_rows = sum(
            1 for objs in all_objects.values()
            for rec in objs if rec.validate()
        )

        lines: list[str] = []
        lines.append(f"Thời gian  : {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
        lines.append(f"File       : {os.path.basename(excel_file)}")
        lines.append(
            f"Tổng       : {total_rows} dòng | "
            f"Lỗi: {total_error_rows} | "
            f"OK: {total_rows - total_error_rows}"
        )

        for sheet_name, objects in all_objects.items():
            # Thu thập lỗi của sheet này trước
            sheet_error_lines: list[str] = []
            for excel_row, record in enumerate(objects, start=2):
                errs = record.validate()
                if errs:
                    sheet_error_lines.append(f"\n  [LOI] Row Excel {excel_row}:")
                    for e in errs:
                        sheet_error_lines.append(f"       - {e}")

            # Header sheet
            err_count = len([l for l in sheet_error_lines if "[LOI]" in l])
            lines.append(f"\n{'=' * 60}")
            lines.append(
                f"  SHEET: {sheet_name}  "
                f"({len(objects)} dòng | {err_count} lỗi)"
            )
            lines.append(f"{'=' * 60}")

            if sheet_error_lines:
                lines.extend(sheet_error_lines)
            else:
                lines.append("  -> Không có lỗi.")

        with open(log_file, "w", encoding="utf-8") as f:
            for line in lines:
                f.write(line + "\n")

        return log_file

    @staticmethod
    def write_excel_summary(excel_file: str, all_objects: dict) -> str:
        """
        Xuất toàn bộ dữ liệu ra Excel, mỗi sheet = 1 XML.
        Thêm 2 cột cuối: STATUS (OK/LOI) và ERRORS (danh sách lỗi).
        Highlight: đỏ nhạt = lỗi, xanh nhạt = OK.
        """
        from openpyxl import Workbook
        from dataclasses import fields as dc_fields
        S         = EXCEL_STYLE
        FILL_HDR  = S.FILL_HDR_SUMMARY
        FILL_ERR  = S.FILL_ROW_ERR
        FILL_OK   = S.FILL_ROW_OK
        FONT_HDR  = S.FONT_HDR
        FONT_DATA = S.FONT_DATA
        FONT_ERR  = S.FONT_ERR
        ALIGN_CTR = S.ALIGN_CTR
        ALIGN_L   = S.ALIGN_L
        BORDER    = S.BORDER

        wb = Workbook()
        if wb.active is not None:
            wb.remove(wb.active)
            

        for sheet_name, objects in all_objects.items():
            if not objects:
                continue
            ws = wb.create_sheet(sheet_name)

            # Lấy tên cột từ dataclass fields (bỏ ClassVar)
            col_names = [f.name for f in dc_fields(objects[0])
                         if not f.name.startswith("_")]
            headers = col_names + ["STATUS", "ERRORS"]

            # Header row
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.fill   = FILL_HDR
                cell.font   = FONT_HDR
                cell.alignment = ALIGN_CTR
                cell.border = BORDER

            # Data rows
            for ri, record in enumerate(objects, start=2):
                errs  = record.validate()
                is_ok = len(errs) == 0

                for ci, fname in enumerate(col_names, 1):
                    val  = getattr(record, fname, None)
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.fill      = FILL_OK if is_ok else FILL_ERR
                    cell.font      = FONT_DATA
                    cell.alignment = ALIGN_L
                    cell.border    = BORDER

                # Cột STATUS
                ci_status = len(col_names) + 1
                cs = ws.cell(row=ri, column=ci_status,
                             value="OK" if is_ok else "LOI")
                cs.fill      = FILL_OK if is_ok else FILL_ERR
                cs.font      = FONT_DATA if is_ok else FONT_ERR
                cs.alignment = ALIGN_CTR
                cs.border    = BORDER

                # Cột ERRORS
                ci_err = len(col_names) + 2
                err_text = "\n".join(f"• {e}" for e in errs) if errs else ""
                ce = ws.cell(row=ri, column=ci_err, value=err_text)
                ce.fill      = FILL_OK if is_ok else FILL_ERR
                ce.font      = FONT_ERR if errs else FONT_DATA
                ce.alignment = ALIGN_L
                ce.border    = BORDER

            # Auto column width (cap 60)
            for ci, h in enumerate(headers, 1):
                max_w = len(h)
                for ri in range(2, ws.max_row + 1):
                    v = ws.cell(row=ri, column=ci).value or ""
                    max_w = max(max_w, min(len(str(v).split("\n")[0]), 60))
                ws.column_dimensions[
                    ws.cell(row=1, column=ci).column_letter
                ].width = max_w + 3

            # Freeze header
            ws.freeze_panes = "A2"
            ws.row_dimensions[1].height = 30

        folder   = os.path.dirname(os.path.abspath(excel_file))
        base     = os.path.splitext(os.path.basename(excel_file))[0]
        now_str  = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_file = os.path.join(folder, f"{base}[{now_str}]_summary.xlsx")
        wb.save(out_file)
        return out_file

    @staticmethod
    def write_excel_errors_only(excel_file: str, all_objects: dict) -> str:
        """
        Xuất chỉ các dòng LỖI ra Excel, mỗi sheet = 1 XML.
        Có thêm cột ROW_EXCEL để tra ngược về file gốc.
        Sheet không có lỗi vẫn xuất header + dòng thông báo sạch.
        """
        from openpyxl import Workbook
        from dataclasses import fields as dc_fields
        S         = EXCEL_STYLE
        FILL_HDR  = S.FILL_HDR_ERRORS
        FILL_ERR  = S.FILL_ROW_ERR
        FILL_OK   = S.FILL_SHEET_CLEAN
        FONT_HDR  = S.FONT_HDR
        FONT_DATA = S.FONT_DATA
        FONT_ERR  = S.FONT_ERR
        ALIGN_CTR = S.ALIGN_CTR
        ALIGN_L   = S.ALIGN_L
        BORDER    = S.BORDER

        wb = Workbook()
        if wb.active is not None:
            wb.remove(wb.active)
        

        for sheet_name, objects in all_objects.items():
            if not objects:
                continue
            ws = wb.create_sheet(sheet_name)

            col_names = [f.name for f in dc_fields(objects[0])
                         if not f.name.startswith("_")]
            headers = ["ROW_EXCEL"] + col_names + ["ERRORS"]

            # Header row
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.fill      = FILL_HDR
                cell.font      = FONT_HDR
                cell.alignment = ALIGN_CTR
                cell.border    = BORDER

            error_rows = 0
            for excel_row, record in enumerate(objects, start=2):
                errs = record.validate()
                if not errs:
                    continue
                error_rows += 1
                write_row = ws.max_row + 1

                # Cột ROW_EXCEL
                cr = ws.cell(row=write_row, column=1, value=excel_row)
                cr.fill = FILL_ERR; cr.font = FONT_DATA
                cr.alignment = ALIGN_CTR; cr.border = BORDER

                # Cột dữ liệu
                for ci, fname in enumerate(col_names, 2):
                    val  = getattr(record, fname, None)
                    cell = ws.cell(row=write_row, column=ci, value=val)
                    cell.fill      = FILL_ERR
                    cell.font      = FONT_DATA
                    cell.alignment = ALIGN_L
                    cell.border    = BORDER

                # Cột ERRORS
                ci_err   = len(col_names) + 2
                err_text = "\n".join(f"• {e}" for e in errs)
                ce = ws.cell(row=write_row, column=ci_err, value=err_text)
                ce.fill      = FILL_ERR
                ce.font      = FONT_ERR
                ce.alignment = ALIGN_L
                ce.border    = BORDER

            # Nếu sheet sạch → ghi 1 dòng thông báo
            if error_rows == 0:
                msg_row = 2
                ws.cell(row=msg_row, column=1,
                        value=f"✓ {sheet_name}: không có lỗi ({len(objects)} dòng)").fill = FILL_OK
                ws.cell(row=msg_row, column=1).font = Font(
                    italic=True, color="2E7D32", name="Arial", size=10)
                ws.merge_cells(
                    start_row=msg_row, start_column=1,
                    end_row=msg_row, end_column=len(headers)
                )

            # Auto column width
            for ci, h in enumerate(headers, 1):
                max_w = len(h)
                for ri in range(2, ws.max_row + 1):
                    v = ws.cell(row=ri, column=ci).value or ""
                    max_w = max(max_w, min(len(str(v).split("\n")[0]), 60))
                ws.column_dimensions[
                    ws.cell(row=1, column=ci).column_letter
                ].width = max_w + 3

            ws.freeze_panes = "B2"  # freeze cả ROW_EXCEL + header
            ws.row_dimensions[1].height = 30

        folder   = os.path.dirname(os.path.abspath(excel_file))
        base     = os.path.splitext(os.path.basename(excel_file))[0]
        now_str  = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_file = os.path.join(folder, f"{base}[{now_str}]_errors_only.xlsx")
        wb.save(out_file)
        return out_file

    @staticmethod
    def print_errors_console(all_objects: dict) -> None:
        """In lỗi ra console theo định dạng dễ đọc."""
        lines = FileHelper._render_summary_lines(all_objects)
        for line in lines:
            print(line)